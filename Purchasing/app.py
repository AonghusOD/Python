from flask import Flask, render_template, request, redirect, Markup
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
app.secret_key = '2515152'
DATA_FILE = 'CorsettiPurchasing.xlsx'

@app.route('/', methods=['GET', 'POST'])
def form():
    if request.method == 'POST':
        name = request.form['name']
        brand = request.form['brand']
        item_model_part = request.form['item_model_part']
        description = request.form['description']
        quantity = request.form['quantity']
        date_needed = request.form['date_needed']
        link = request.form['link']
        brand_substitute_ok = request.form.get('brand_substitute_ok', False)
        model_substitute_ok = request.form.get('model_substitute_ok', False)
        ordered = request.form['ordered']

        if brand_substitute_ok:
            brand_substitute_ok = "Yes"
        else:
            brand_substitute_ok = "No"

        if model_substitute_ok:
            model_substitute_ok = "Yes"
        else:
            model_substitute_ok = "No"

        # Save data to Excel file
        if not os.path.exists(DATA_FILE):
            wb = Workbook()
            wb.save(DATA_FILE)

        wb = load_workbook(DATA_FILE)
        sheet = wb.active
        sheet.append([name, brand, item_model_part, description, quantity, date_needed, link, brand_substitute_ok, model_substitute_ok, ordered])
        wb.save(DATA_FILE)

        return redirect('/')

    else:
        # Load data from Excel file
        if os.path.exists(DATA_FILE):
            wb = load_workbook(DATA_FILE)
            sheet = wb.active
            data = sheet.iter_rows(values_only=True)
        else:
            data = []

        modified_data = []
        for row in data:
            modified_row = list(row)
            link_value = row[6]  # Assuming the link column is at index 6
            hyperlink = Markup(f'<a href="{link_value}">{link_value}</a>')
            modified_row[6] = hyperlink
            modified_data.append(modified_row)

        return render_template('form.html', data=modified_data)


@app.route('/delete/<int:index>', methods=['GET', 'POST'])
def delete(index):
    if os.path.exists(DATA_FILE):
        wb = load_workbook(DATA_FILE)
        sheet = wb.active
        sheet.delete_rows(index + 1, 1)  # Adjust for header row
        wb.save(DATA_FILE)

    return redirect('/')


@app.route('/toggle/<int:index>', methods=['GET', 'POST'])
def toggle_ordered(index):
    if os.path.exists(DATA_FILE):
        wb = load_workbook(DATA_FILE)
        sheet = wb.active

        cell = sheet.cell(row=index + 1, column=10)  # Adjust for header row
        current_value = cell.value
        new_value = "No" if current_value == "Yes" else "Yes"
        cell.value = new_value
        wb.save(DATA_FILE)

    return redirect('/')


@app.route('/favicon.ico')
def fav():
    return send_from_directory(os.path.join(app.root_path, 'static'), 'css.ico')


if __name__ == '__main__':
    app.run()
